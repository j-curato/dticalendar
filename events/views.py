# from rest_framework.response import Response
# from rest_framework.decorators import api_view
# from .serializers import EventSerializer
from django.http import HttpResponse
from django.http import JsonResponse
from django.db import connection
from django.db.models import F
from django.db.models.functions import Lower
from django.db import connection
from calendars.models import Calendar
from divisions.models import Division
from units.models import Unit
from orgoutcomes.models import OrgOutcome
from paps.models import Pap
from provinces.models import Province
from lgus.models import Lgu
from barangays.models import Barangay
from django.shortcuts import get_object_or_404
from django.db.models import Q
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.shortcuts import render
from django.http import FileResponse
from datetime import datetime
from mimetypes import guess_type
from .models import Event
from offices.models import Office
from django.contrib.auth.decorators import login_required
import os
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
# make_random_password is a method from django.contrib.auth.models

# @api_view(['GET'])
# def get_events(request):
#     events = Event.objects.all()
#     serializer = EventSerializer(events, many=True)
#     return Response(serializer.data)
@csrf_exempt
def save_event_ajax(request):

    if request.method == 'POST':
        # ... (existing code to retrieve IDs)
        calendar_id = request.POST.get('calendar_id')
        division_id = request.POST.get('division_id')
        orgoutcome_id = request.POST.get('orgoutcome_id')
        pap_id = request.POST.get('pap_id')
        province_id = request.POST.get('province_id')
        lgu_id = request.POST.get('lgu_id')
        barangay_id = request.POST.get('barangay_id')

        # Check if the user is logged in
        if request.user:
            calendar = get_object_or_404(Calendar, pk=calendar_id)
            division = get_object_or_404(Division, pk=division_id)
            orgoutcome = get_object_or_404(OrgOutcome, pk=orgoutcome_id)
            pap = get_object_or_404(Pap, pk=pap_id)
            province = get_object_or_404(Province, pk=province_id)
            lgu = get_object_or_404(Lgu, pk=lgu_id)
            barangay = get_object_or_404(Barangay, pk=barangay_id) if barangay_id and str(barangay_id) != '0' else None

        start_date = timezone.datetime.strptime(request.POST['whole_date_start'], "%Y-%m-%dT%H:%M")
        end_date = timezone.datetime.strptime(request.POST['whole_date_end'], "%Y-%m-%dT%H:%M")

        # Fetch existing events based on event_code and date range
        event_code = request.POST['event_code']
        existing_event = Event.objects.filter(event_code=event_code, whole_date_start__range=(start_date, end_date))

        # Update existing event
        updated_fields = {}

        # ... (check and update fields)
        if existing_event.whole_date_start_searchable != start_date.strftime("%B %d, %Y"):
            existing_event.whole_date_start_searchable = start_date.strftime("%B %d, %Y")
            updated_fields['whole_date_start_searchable'] = existing_event.whole_date_start_searchable
        
        if existing_event.event_title != request.POST['event_title'].upper():
            existing_event.event_title = request.POST['event_title'].upper()
            updated_fields['event_title'] = existing_event.event_title

        if existing_event.event_desc != request.POST['event_desc'].upper():
            existing_event.event_desc = request.POST['event_desc'].upper()
            updated_fields['event_desc'] = existing_event.event_desc

        if existing_event.participants != request.POST['participants'].upper():
            existing_event.participants = request.POST['participants'].upper()
            updated_fields['participants'] = existing_event.participants

        if existing_event.event_location != request.POST['event_location'].upper():
            existing_event.event_location = request.POST['event_location'].upper()
            updated_fields['event_location'] = existing_event.event_location

        if existing_event.event_day_start != start_date.day:
            existing_event.event_day_start = start_date.day
            updated_fields['event_day_start'] = existing_event.event_day_start

        if existing_event.event_month_start != start_date.month:
            existing_event.event_month_start = start_date.month
            updated_fields['event_month_start'] = existing_event.event_month_start

        if existing_event.event_year_start != start_date.year:
            existing_event.event_year_start = start_date.year
            updated_fields['event_month_start'] = existing_event.event_year_start

        if existing_event.event_time_start != request.POST['event_time_start']:
            existing_event.event_time_start = request.POST['event_time_start']
            updated_fields['event_time_start'] = existing_event.event_time_start

        if existing_event.event_day_end != request.POST['event_day_end']:
            existing_event.event_day_end = request.POST['event_day_end']
            updated_fields['event_day_end'] = existing_event.event_day_end

        if existing_event.event_month_end != request.POST['event_month_end']:
            existing_event.event_month_end = request.POST['event_month_end']
            updated_fields['event_month_end'] = existing_event.event_month_end

        if existing_event.event_year_end != request.POST['event_year_end']:
            existing_event.event_year_end = request.POST['event_year_end']
            updated_fields['event_year_end'] = existing_event.event_year_end

        if existing_event.event_time_end != request.POST['event_time_end']:
            existing_event.event_time_end = request.POST['event_time_end']
            updated_fields['event_time_end'] = existing_event.event_time_end

        if existing_event.whole_date_start != start_date:
            existing_event.whole_date_start = start_date
            updated_fields['whole_date_start'] = existing_event.whole_date_start

        if existing_event.whole_date_end != end_date:
            existing_event.whole_date_end = end_date
            updated_fields['whole_date_end'] = existing_event.whole_date_end

        if existing_event.division_id != division:
            existing_event.division_id = division
            updated_fields['division_id'] = existing_event.division_id

        if existing_event.orgoutcome_id != orgoutcome:
            existing_event.orgoutcome_id = orgoutcome
            updated_fields['orgoutcome_id'] = existing_event.orgoutcome_id

        if existing_event.pap_id != pap:
            existing_event.pap_id = pap
            updated_fields['pap_id'] = existing_event.pap_id

        if existing_event.province_id != province:
            existing_event.province_id = province
            updated_fields['province_id'] = existing_event.province_id

        if existing_event.barangay_id != barangay:
            existing_event.barangay_id = barangay
            updated_fields['barangay_id'] = existing_event.barangay_id

        #filename = os.path.basename(existing_event.file_attachment)
        if os.path.basename(existing_event.file_attachment) != request.FILES.get('file_attachment'):
            existing_event.file_attachment = request.FILES.get('file_attachment')
            updated_fields['file_attachment'] = existing_event.file_attachment

        if existing_event.whole_date_start_searchable != start_date.strftime("%B %d, %Y"):
            existing_event.whole_date_start_searchable = start_date.strftime("%B %d, %Y")
            updated_fields['whole_date_start_searchable'] = existing_event.whole_date_start_searchable

        if existing_event.whole_date_end_searchable != request.POST['whole_date_end_searchable']:
            existing_event.whole_date_end_searchable = request.POST['whole_date_end_searchable']
            updated_fields['whole_date_end_searchable'] = existing_event.whole_date_end_searchable

        if existing_event.office != request.POST['office'].upper():
            existing_event.office = request.POST['office'].upper()
            updated_fields['office'] = existing_event.office

        if existing_event.org_outcome != request.POST['org_outcome'].upper():
            existing_event.org_outcome = request.POST['org_outcome'].upper()
            updated_fields['org_outcome'] = existing_event.org_outcome

        if existing_event.paps != request.POST['paps'].upper():
            existing_event.paps = request.POST['paps'].upper()
            updated_fields['paps'] = existing_event.paps

        if existing_event.unit != request.POST['unit'].upper():
            existing_event.unit = request.POST['unit'].upper()
            updated_fields['unit'] = existing_event.unit

        if existing_event.division_name != request.POST['division_name'].upper():
            existing_event.division_name = request.POST['division_name'].upper()
            updated_fields['division_name'] = existing_event.division_name

        if existing_event.whole_dateStart_with_time != timezone.datetime.strptime(start_date.strftime("%Y-%m-%dT%H:%M"), "%Y-%m-%dT%H:%M"):
            existing_event.whole_dateStart_with_time = timezone.datetime.strptime(start_date.strftime("%Y-%m-%dT%H:%M"), "%Y-%m-%dT%H:%M")
            updated_fields['whole_dateStart_with_time'] = existing_event.whole_dateStart_with_time

        if existing_event.whole_dateEnd_with_time != timezone.datetime.strptime(request.POST['whole_date_end'], "%Y-%m-%dT%H:%M"):
            existing_event.whole_dateEnd_with_time = timezone.datetime.strptime(request.POST['whole_date_end'], "%Y-%m-%dT%H:%M")
            updated_fields['whole_dateEnd_with_time'] = existing_event.whole_dateEnd_with_time

        if existing_event.event_location_district != request.POST['event_location_district'].upper():
            existing_event.event_location_district = request.POST['event_location_district'].upper()
            updated_fields['event_location_district'] = existing_event.event_location_district

        if existing_event.event_location_lgu != request.POST['event_location_lgu'].upper():
            existing_event.event_location_lgu = request.POST['event_location_lgu'].upper()
            updated_fields['event_location_lgu'] = existing_event.event_location_lgu

        if existing_event.event_location_barangay != request.POST['event_location_barangay'].upper():
            existing_event.event_location_barangay = request.POST['event_location_barangay'].upper()
            updated_fields['event_location_barangay'] = existing_event.event_location_barangay

        if existing_event.event_all_day != request.POST['event_all_day']:
            existing_event.event_all_day = request.POST['event_all_day']
            updated_fields['event_all_day'] = existing_event.event_all_day

        # Use the update method to apply changes to all selected records in a single database query
        existing_event.update(**updated_fields)

            # Create a new event if no existing event with the specified event_code
        if request.POST['buttontxt'] == 'Save':

            event = Event()
            # assign to event.user the current logged-in user id
            event.user = request.user
            event.event_title = request.POST['event_title'].upper()
            event.event_desc = request.POST['event_desc'].upper()
            event.participants = request.POST['participants'].upper()
            event.event_location = request.POST['event_location'].upper()
            event.event_day_start = start_date.day
            event.event_month_start = start_date.month
            event.event_year_start = start_date.year
            event.event_time_start = request.POST['event_time_start']
            event.event_day_end = request.POST['event_day_end']
            event.event_month_end = request.POST['event_month_end']
            event.event_year_end = request.POST['event_year_end']
            event.event_time_end = request.POST['event_time_end']
            event.whole_date_start = start_date
            event.whole_date_end = end_date
            event.calendar = calendar
            event.division = division
            event.orgoutcome = orgoutcome
            event.pap = pap
            event.province = province
            event.lgu = lgu
            event.barangay = barangay
            # check if the file_attachment field is not empty
            if request.FILES.get('file_attachment') != None:
                event.file_attachment = request.FILES.get('file_attachment')
            else:
                event.file_attachment = 'NONE'
            
            #event.file_attachment = request.FILES.get('file_attachment')
            event.whole_date_start_searchable = start_date.strftime("%B %d, %Y")
            event.whole_date_end_searchable = request.POST['whole_date_end_searchable']
            event.office = request.POST['office'].upper()
            event.org_outcome = request.POST['org_outcome'].upper()
            event.paps = request.POST['paps'].upper()
            event.unit = request.POST['unit'].upper()
            event.division_name = request.POST['division_name'].upper()
            #event.whole_dateStart_with_time = timezone.datetime.strptime(request.POST['whole_date_start'], "%Y-%m-%dT%H:%M")
            event.whole_dateStart_with_time = timezone.datetime.strptime(start_date.strftime("%Y-%m-%dT%H:%M"), "%Y-%m-%dT%H:%M")
            event.whole_dateEnd_with_time = timezone.datetime.strptime(request.POST['whole_date_end'], "%Y-%m-%dT%H:%M")
            event.actual_outcome = "PENDING".upper()
            event.calendar_name = request.POST['calendar_name'].upper()
            event.event_location_district = request.POST['event_location_district'].upper()
            event.event_location_lgu = request.POST['event_location_lgu'].upper()
            event.event_location_barangay = request.POST['event_location_barangay'].upper()
            event.event_status = "PENDING".upper()
            event.expected_outcome = "PENDING".upper()
            event.event_code = request.POST['event_code']
            # save a boolean value to event_all_day field
            if request.POST['event_all_day'] == 'true':
                event.event_all_day = True
            else:
                event.event_all_day = False

            event.save()

        return JsonResponse({'message': 'True'})
    else:
        return JsonResponse({'message': 'False'})

# save using ajax
@csrf_exempt
def save_event_ajax_version1(request):
    
    if request.method == 'POST':
        calendar_id = request.POST.get('calendar_id')
        division_id = request.POST.get('division_id')
        orgoutcome_id = request.POST.get('orgoutcome_id')
        pap_id = request.POST.get('pap_id')
        province_id = request.POST.get('province_id')
        lgu_id = request.POST.get('lgu_id')
        barangay_id = request.POST.get('barangay_id')

        # check if the user is logged in
        if request.user:
             calendar = get_object_or_404(Calendar, pk=calendar_id)
             division = get_object_or_404(Division, pk=division_id)
             orgoutcome = get_object_or_404(OrgOutcome, pk=orgoutcome_id)
             pap = get_object_or_404(Pap, pk=pap_id)
             province = get_object_or_404(Province, pk=province_id)
             lgu = get_object_or_404(Lgu, pk=lgu_id)
             barangay = get_object_or_404(Barangay, pk=barangay_id) if barangay_id and str(barangay_id) != '0' else None

        # set variables to use for multiple records to be saved if the event is recurring
        start_date = timezone.datetime.strptime(request.POST['whole_date_start'], "%Y-%m-%dT%H:%M")
        end_date = timezone.datetime.strptime(request.POST['whole_date_end'], "%Y-%m-%dT%H:%M")
        day_duration = timezone.timedelta(days=1)

        # use while loop to save multiple records if the event is recurring
        while start_date <= end_date:

            event = Event()
            # assign to event.user the current logged-in user id
            event.user = request.user
            event.event_title = request.POST['event_title'].upper()
            event.event_desc = request.POST['event_desc'].upper()
            event.participants = request.POST['participants'].upper()
            event.event_location = request.POST['event_location'].upper()
            event.event_day_start = start_date.day
            event.event_month_start = start_date.month
            event.event_year_start = start_date.year
            event.event_time_start = request.POST['event_time_start']
            event.event_day_end = request.POST['event_day_end']
            event.event_month_end = request.POST['event_month_end']
            event.event_year_end = request.POST['event_year_end']
            event.event_time_end = request.POST['event_time_end']
            event.whole_date_start = start_date
            event.whole_date_end = end_date
            event.calendar = calendar
            event.division = division
            event.orgoutcome = orgoutcome
            event.pap = pap
            event.province = province
            event.lgu = lgu
            event.barangay = barangay
            # check if the file_attachment field is not empty
            if request.FILES.get('file_attachment') != None:
                event.file_attachment = request.FILES.get('file_attachment')
            else:
                event.file_attachment = 'NONE'
            
            event.whole_date_start_searchable = start_date.strftime("%B %d, %Y")
            event.whole_date_end_searchable = request.POST['whole_date_end_searchable']
            event.office = request.POST['office'].upper()
            event.org_outcome = request.POST['org_outcome'].upper()
            event.paps = request.POST['paps'].upper()
            event.unit = request.POST['unit'].upper()
            event.division_name = request.POST['division_name'].upper()
            event.whole_dateStart_with_time = timezone.datetime.strptime(start_date.strftime("%Y-%m-%dT%H:%M"), "%Y-%m-%dT%H:%M")
            event.whole_dateEnd_with_time = timezone.datetime.strptime(request.POST['whole_date_end'], "%Y-%m-%dT%H:%M")
            event.actual_outcome = "PENDING".upper()
            event.calendar_name = request.POST['calendar_name'].upper()
            event.event_location_district = request.POST['event_location_district'].upper()
            event.event_location_lgu = request.POST['event_location_lgu'].upper()
            event.event_location_barangay = request.POST['event_location_barangay'].upper()
            event.event_status = "PENDING".upper()
            event.expected_outcome = "PENDING".upper()
            event.event_code = request.POST['event_code']
            # save a boolean value to event_all_day field
            if request.POST['event_all_day'] == 'true':
                event.event_all_day = True
            else:
                event.event_all_day = False

            event.save()
            start_date = start_date + day_duration

        return JsonResponse({'message': 'True'})
    else:
        return JsonResponse({'message': 'False'})
    

@csrf_exempt
def save_event_ajax_ver2(request):
    
    if request.method == 'POST':

        calendar_id = request.POST.get('calendar_id')
        division_id = request.POST.get('division_id')
        orgoutcome_id = request.POST.get('orgoutcome_id')
        pap_id = request.POST.get('pap_id')
        province_id = request.POST.get('province_id')
        lgu_id = request.POST.get('lgu_id')
        barangay_id = request.POST.get('barangay_id')
        unit_id = request.POST.get('unit')
       

        # check if the user is logged in
        if request.user:
             calendar = get_object_or_404(Calendar, pk=calendar_id)
             division = get_object_or_404(Division, pk=division_id)
             orgoutcome = get_object_or_404(OrgOutcome, pk=orgoutcome_id)
             pap = get_object_or_404(Pap, pk=pap_id)
             province = get_object_or_404(Province, pk=province_id)
             lgu = get_object_or_404(Lgu, pk=lgu_id)
             barangay = get_object_or_404(Barangay, pk=barangay_id) if barangay_id and str(barangay_id) != '0' else None
             unit = get_object_or_404(Unit, pk=unit_id)

        # set variables to use for multiple records to be saved if the event is recurring
        start_date = timezone.datetime.strptime(request.POST['whole_date_start'], "%Y-%m-%dT%H:%M")
        end_date = timezone.datetime.strptime(request.POST['whole_date_end'], "%Y-%m-%dT%H:%M")
        #day_duration = timezone.timedelta(days=1)
     
    if request.POST['buttontxt'] == 'Save':
        
        event = Event()
        # assign to event.user the current logged-in user id
        event.user = request.user
        event.event_title = request.POST['event_title'].upper()
        event.event_desc = request.POST['event_desc'].upper()
        event.participants = request.POST['participants'].upper()
        event.event_location = request.POST['event_location'].upper()
        event.event_day_start = start_date.day
        event.event_month_start = start_date.month
        event.event_year_start = start_date.year
        event.event_time_start = request.POST['event_time_start']
        event.event_day_end = request.POST['event_day_end']
        event.event_month_end = request.POST['event_month_end']
        event.event_year_end = request.POST['event_year_end']
        event.event_time_end = request.POST['event_time_end']
        event.whole_date_start = start_date
        event.whole_date_end = end_date
        event.calendar = calendar
        event.division = division
        event.orgoutcome = orgoutcome
        event.pap = pap
        event.province = province
        event.lgu = lgu
        event.barangay = barangay
        # check if the file_attachment field is not empty
        if request.FILES.get('file_attachment') != None:
            event.file_attachment = request.FILES.get('file_attachment')
        else:
            event.file_attachment = 'NONE'
        
        event.whole_date_start_searchable = start_date.strftime("%B %d, %Y")
        event.whole_date_end_searchable = request.POST['whole_date_end_searchable']
        event.office = request.POST['office'].upper()
        fk_office_id = request.POST.get('fk_office_id')
        if fk_office_id:
            event.fk_office_id = fk_office_id
        event.org_outcome = request.POST['org_outcome'].upper()
        event.paps = request.POST['paps'].upper()
        event.unit = unit
        event.unit_name = request.POST['unit_name'].upper()
        event.division_name = request.POST['division_name'].upper()
        event.whole_dateStart_with_time = timezone.datetime.strptime(start_date.strftime("%Y-%m-%dT%H:%M"), "%Y-%m-%dT%H:%M")
        event.whole_dateEnd_with_time = timezone.datetime.strptime(request.POST['whole_date_end'], "%Y-%m-%dT%H:%M")
        event.actual_outcome = "PENDING".upper()
        event.calendar_name = request.POST['calendar_name'].upper()
        event.event_location_district = request.POST['event_location_district'].upper()
        event.event_location_lgu = request.POST['event_location_lgu'].upper()
        event.event_location_barangay = request.POST['event_location_barangay'].upper()
        event.event_status = "PENDING".upper()
        event.expected_outcome = "PENDING".upper()
        event.event_code = request.POST['event_code']
        # save a boolean value to event_all_day field
        if request.POST['event_all_day'] == 'true':
            event.event_all_day = True
        else:
            event.event_all_day = False

        if event.save():
            return JsonResponse({'message': 'True'})
        else:
            return JsonResponse({'message': 'False'})
            
    else:

        evenpID = request.POST['pID']
        existing_event = Event.objects.get(pk=evenpID)
        updated_fields = {}

        if existing_event.event_title != request.POST['event_title'].upper():
            existing_event.event_title = request.POST['event_title'].upper()
            updated_fields['event_title'] = existing_event.event_title

        if existing_event.event_desc != request.POST['event_desc'].upper():
            existing_event.event_desc = request.POST['event_desc'].upper()
            updated_fields['event_desc'] = existing_event.event_desc

        if existing_event.participants != request.POST['participants'].upper():
            existing_event.participants = request.POST['participants'].upper()
            updated_fields['participants'] = existing_event.participants

        if existing_event.event_location != request.POST['event_location'].upper():
            existing_event.event_location = request.POST['event_location'].upper()
            updated_fields['event_location'] = existing_event.event_location

        if existing_event.event_day_start != start_date.day:
            existing_event.event_day_start = start_date.day
            updated_fields['event_day_start'] = existing_event.event_day_start

        if existing_event.event_month_start != start_date.month:
            existing_event.event_month_start = start_date.month
            updated_fields['event_month_start'] = existing_event.event_month_start

        if existing_event.event_year_start != start_date.year:
            existing_event.event_year_start = start_date.year
            updated_fields['event_month_start'] = existing_event.event_year_start

        if existing_event.event_time_start != request.POST['event_time_start']:
            existing_event.event_time_start = request.POST['event_time_start']
            updated_fields['event_time_start'] = existing_event.event_time_start

        if existing_event.event_day_end != request.POST['event_day_end']:
            existing_event.event_day_end = request.POST['event_day_end']
            updated_fields['event_day_end'] = existing_event.event_day_end

        if existing_event.event_month_end != request.POST['event_month_end']:
            existing_event.event_month_end = request.POST['event_month_end']
            updated_fields['event_month_end'] = existing_event.event_month_end

        if existing_event.event_year_end != request.POST['event_year_end']:
            existing_event.event_year_end = request.POST['event_year_end']
            updated_fields['event_year_end'] = existing_event.event_year_end

        if existing_event.event_time_end != request.POST['event_time_end']:
            existing_event.event_time_end = request.POST['event_time_end']
            updated_fields['event_time_end'] = existing_event.event_time_end

        if existing_event.whole_date_start != start_date:
            existing_event.whole_date_start = start_date
            updated_fields['whole_date_start'] = existing_event.whole_date_start

        if existing_event.whole_date_end != end_date:
            existing_event.whole_date_end = end_date
            updated_fields['whole_date_end'] = existing_event.whole_date_end

        if existing_event.division_id != division:
            existing_event.division_id = division
            updated_fields['division_id'] = existing_event.division_id

        if existing_event.unit_id != unit:
            existing_event.unit_id = unit
            updated_fields['unit_id'] = existing_event.unit_id

        if existing_event.orgoutcome_id != orgoutcome:
            existing_event.orgoutcome_id = orgoutcome
            updated_fields['orgoutcome_id'] = existing_event.orgoutcome_id

        if existing_event.pap_id != pap:
            existing_event.pap_id = pap
            updated_fields['pap_id'] = existing_event.pap_id

        if existing_event.province_id != province:
            existing_event.province_id = province
            updated_fields['province_id'] = existing_event.province_id

        if existing_event.lgu_id != request.POST['lgu_id']:
            existing_event.lgu_id = request.POST['lgu_id']
            updated_fields['lgu_id'] =  existing_event.lgu_id

        if existing_event.barangay_id != barangay:
            existing_event.barangay_id = barangay
            updated_fields['barangay_id'] = existing_event.barangay_id

        if existing_event.whole_date_start_searchable != start_date.strftime("%B %d, %Y"):
            existing_event.whole_date_start_searchable = start_date.strftime("%B %d, %Y")
            updated_fields['whole_date_start_searchable'] = existing_event.whole_date_start_searchable

        if existing_event.whole_date_end_searchable != request.POST['whole_date_end_searchable']:
            existing_event.whole_date_end_searchable = request.POST['whole_date_end_searchable']
            updated_fields['whole_date_end_searchable'] = existing_event.whole_date_end_searchable

        if existing_event.office != request.POST['office'].upper():
            existing_event.office = request.POST['office'].upper()
            updated_fields['office'] = existing_event.office

        fk_office_id = request.POST.get('fk_office_id')
        if fk_office_id and str(existing_event.fk_office_id) != str(fk_office_id):
            existing_event.fk_office_id = fk_office_id
            updated_fields['fk_office_id'] = fk_office_id

        if existing_event.org_outcome != request.POST['org_outcome'].upper():
            existing_event.org_outcome = request.POST['org_outcome'].upper()
            updated_fields['org_outcome'] = existing_event.org_outcome

        if existing_event.paps != request.POST['paps'].upper():
            existing_event.paps = request.POST['paps'].upper()
            updated_fields['paps'] = existing_event.paps

        if existing_event.unit_name != request.POST['unit_name'].upper():
            existing_event.unit_name = request.POST['unit_name'].upper()
            updated_fields['unit_name'] = existing_event.unit_name

        if existing_event.division_name != request.POST['division_name'].upper():
            existing_event.division_name = request.POST['division_name'].upper()
            updated_fields['division_name'] = existing_event.division_name

        if existing_event.whole_dateStart_with_time != timezone.datetime.strptime(start_date.strftime("%Y-%m-%dT%H:%M"), "%Y-%m-%dT%H:%M"):
            existing_event.whole_dateStart_with_time = timezone.datetime.strptime(start_date.strftime("%Y-%m-%dT%H:%M"), "%Y-%m-%dT%H:%M")
            updated_fields['whole_dateStart_with_time'] = existing_event.whole_dateStart_with_time

        if existing_event.whole_dateEnd_with_time != timezone.datetime.strptime(request.POST['whole_date_end'], "%Y-%m-%dT%H:%M"):
            existing_event.whole_dateEnd_with_time = timezone.datetime.strptime(request.POST['whole_date_end'], "%Y-%m-%dT%H:%M")
            updated_fields['whole_dateEnd_with_time'] = existing_event.whole_dateEnd_with_time

        if existing_event.event_location_district != request.POST['event_location_district'].upper():
            existing_event.event_location_district = request.POST['event_location_district'].upper()
            updated_fields['event_location_district'] = existing_event.event_location_district

        if existing_event.event_location_lgu != request.POST['event_location_lgu'].upper():
            existing_event.event_location_lgu = request.POST['event_location_lgu'].upper()
            updated_fields['event_location_lgu'] = existing_event.event_location_lgu

        if existing_event.event_location_barangay != request.POST['event_location_barangay'].upper():
            existing_event.event_location_barangay = request.POST['event_location_barangay'].upper()
            updated_fields['event_location_barangay'] = existing_event.event_location_barangay

        # Check and handle boolean field
        requested_event_all_day = request.POST['event_all_day'].lower() == 'true'
        if existing_event.event_all_day != requested_event_all_day:
            existing_event.event_all_day = requested_event_all_day
            updated_fields['event_all_day'] = existing_event.event_all_day

        if 'file_attachment' in request.FILES:
            new_file_attachment = request.FILES['file_attachment']
            # Check if the new file is different from the existing one
            if existing_event.file_attachment != new_file_attachment:
                existing_event.file_attachment = new_file_attachment
                updated_fields['file_attachment'] = existing_event.file_attachment

        if updated_fields:
            existing_event.save()  # Save the changes to the database
            return JsonResponse({'message': 'True'})
        else:
            return JsonResponse({'message': 'False'})


    # method to fetch the event table data and display it in the following field order - whole_date_start_searchable, event_title, office.
    # but the office data is divided into 5 fields that is based on its values. Namely: RO, ADN, ADS, SDN, SDS and PDI
    # display the data in event_display.html

def get_eventsList(request):
    txturl = 'events'
    events = Event.objects.all()
    return render(request, 'events/event_display.html', {'events': events, 'txturl': txturl})

@csrf_exempt
def get_eventsListDate(request):
    # check request method
    if request.method != 'POST':
        return JsonResponse({'error': 'POST request required.'}, status=400)
    # Extract start_date and end_date from request body
    try:
        request_data = json.loads(request.body)
        start_date = request_data.get('start_date')
        end_date = request_data.get('end_date')
    except:
        return JsonResponse({'error': 'Invalid request body.'}, status=400)
    
    # define the desired fields to include in the JSON response
    fields_to_include = ['id','event_title', 'event_desc', 'participants', 'event_location', 'whole_date_start_searchable', 'whole_date_end_searchable', 'event_time_start', 'event_time_end', 'office', 'org_outcome', 'paps', 'unit_name', 'division_name', 'actual_outcome', 'event_location_lgu', 'event_location_barangay', 'event_status', 'expected_outcome','event_all_day']
    # Filter events based on start_date and end_date
    events = Event.objects.filter(Q(whole_date_start__gte=start_date) & Q(whole_date_start__lte=end_date))
    events_json = []
    # Loop through each event and extract the desired fields
    for event in events:
        event_json = {}
        for field in fields_to_include:
            event_json[field] = getattr(event, field)
        events_json.append(event_json)
    # Return the JSON response
    return JsonResponse(events_json, safe=False)

@csrf_exempt
def get_event_details(request):
    # check request method
    if request.method != 'POST':
        return JsonResponse({'error': 'POST request required.'}, status=400)
    # extract id from the request body
    try:
        request_data = json.loads(request.body)
        id = request_data.get('event_id')
    except:
        return JsonResponse({'error': 'Invalid request body.'}, status=400)
    
    # filter event based on id
    event = Event.objects.filter(id=id)
    # return the JSON response
    return JsonResponse(list(event.values()), safe=False)

# method to display event details using datatable server side processing
def fetch_events_ajax(request):
    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '')
        order_direction = request.GET.get('order[0][dir]', 'asc')
        sort_dir = 'DESC' if order_direction == 'desc' else 'ASC'
        year = int(request.GET.get('year', timezone.now().year))

        # Fetch offices dynamically from tbl_offices
        offices = list(Office.objects.order_by('id').values_list('office_initials', flat=True))

        # Define searchable columns dynamically
        columns = ['whole_date_start_searchable'] + offices

        # Build dynamic CASE WHEN pivot for each office
        case_statements = ',\n            '.join([
            f"MAX(CASE WHEN office = '{o}' THEN event_titles END) AS \"{o}\""
            for o in offices
        ])

        query = f"""
        SELECT
            DISTINCT generated_date,
            TO_CHAR(generated_date, 'FMMonth DD, YYYY') AS whole_date_start_searchable,
            {case_statements}
        FROM (
            SELECT
                generated_date,
                office,
                STRING_AGG(CONCAT(event_title, '*', id, '*', division_name, '*', unit_name, '*', event_time_start, '*', event_time_end), '|||') AS event_titles
            FROM (
                SELECT
                    generate_series(whole_date_start::date, whole_date_end::date, '1 day'::interval)::date AS generated_date,
                    office,
                    event_title,
                    id,
                    division_name,
                    unit_name,
                    event_time_start,
                    event_time_end
                FROM events_event
                WHERE display_status = true
            ) AS date_series
            WHERE EXTRACT(YEAR FROM generated_date) = {year}
            GROUP BY generated_date, office
        ) AS subquery
        GROUP BY generated_date
        ORDER BY generated_date {sort_dir};
        """

        with connection.cursor() as cursor:
            cursor.execute(query)
            result = cursor.fetchall()

        # Convert the result into a list of dictionaries dynamically
        data = []
        for row in result:
            entry = {
                'whole_date_start': row[0].strftime('%Y-%m-%d'),
                'whole_date_start_searchable': row[1],
            }
            for i, office in enumerate(offices):
                entry[office] = row[2 + i]
            data.append(entry)

        # Apply the search filter dynamically
        filtered_data = [
            entry for entry in data if any(
                search_value.lower() in (str(entry[col]).lower() if entry[col] is not None else '') for col in columns
            )
        ]

        response_data = {
            'draw': draw,
            'recordsTotal': len(data),
            'recordsFiltered': len(filtered_data),
            'data': filtered_data[start:start + length]
        }

        return JsonResponse(response_data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
# load division datatables html page - events_display_div.html
def load_div_datatbl_html(request):
    txturl = 'events'
    officetxt = request.GET.get('dtioffice')
    divisions = Division.objects.all()
    return render(request, 'events/events_display_div.html', {'divisions': divisions, 'txturl': txturl, 'officetxt': officetxt})

def load_unit_datatbl_html(request):
    txturl = 'events'
    officetxt = request.GET.get('dtiofficeparam')
    divtxt = request.GET.get('dtidivision')
    divisions = Division.objects.all()
    return render(request, 'events/events_display_unit.html', {'divisions': divisions, 'txturl': txturl, 'officetxt': officetxt, 'divtxt': divtxt})

def fetch_events_by_unit_ajax_ver1(request):
    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '')
        #office_txt = request.GET.get('office', '')

        # Define the columns you want to search on
        # columns = ['whole_date_start_searchable']
        columns = ['whole_date_start_searchable', 'NC', 'EDU', 'TPU', 'CARP', 'GAD', 'PLANNING']

        query = """
        SELECT whole_date_start,
            whole_date_start_searchable, 
            MAX(CASE WHEN unit = 'NC' THEN event_titles END) AS NC,
            MAX(CASE WHEN unit = 'EDU' THEN event_titles END) AS EDU,
            MAX(CASE WHEN unit = 'TPU' THEN event_titles END) AS TPU,
            MAX(CASE WHEN unit = 'CARP' THEN event_titles END) AS CARP,
            MAX(CASE WHEN unit = 'GAD' THEN event_titles END) AS GAD,
            MAX(CASE WHEN unit = 'PLANNING' THEN event_titles END) AS PLANNING
        FROM (
            SELECT whole_date_start,
                whole_date_start_searchable,
                unit,
                STRING_AGG(event_title, ', ') AS event_titles, division_name, office
            FROM events_event
            WHERE division_name = %s AND office = %s
            GROUP BY whole_date_start, whole_date_start_searchable, unit, division_name, office
        ) AS subquery
        GROUP BY whole_date_start, whole_date_start_searchable
        ORDER BY 1;
        """

        with connection.cursor() as cursor:
            cursor.execute(query, [request.GET.get('division'), request.GET.get('office')])
            result = cursor.fetchall()

        # Convert the result into a list of dictionaries
        data = []
        for row in result:
            data.append({
                'whole_date_start': row[0].strftime('%Y-%m-%d'),  # Format as needed
                'whole_date_start_searchable': row[1],
                'NC': row[2],
                'EDU': row[3],
                'TPU': row[4],
                'CARP': row[5],
                'GAD': row[6],
                'PLANNING': row[7]
                # Add more fields as needed
            })
            
        # Apply the search filter to the data
        # filtered_data = [entry for entry in data if any(search_value.lower() in entry[col].lower() for col in columns)]
        filtered_data = [
            entry for entry in data if any(
                search_value.lower() in (str(entry[col]).lower() if entry[col] is not None else '') for col in columns
            )
        ]

        response_data = {
            'draw': draw,
            'recordsTotal': len(data),
            'recordsFiltered': len(filtered_data),
            'data': filtered_data[start:start + length]
        }

        return JsonResponse(response_data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    

def fetch_events_by_div_ajax(request):
    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '')
        year = int(request.GET.get('year', timezone.now().year))

        # Fetch division names filtered by the selected office
        office = request.GET.get('office', '')
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT dd.division_name
                FROM divisions_division dd
                INNER JOIN tbl_offices o ON dd.fk_office_id = o.id
                WHERE o.office_initials = %s
                ORDER BY dd.id
            """, [office])
            division_names = [row[0] for row in cursor.fetchall()]

        if not division_names:
            return JsonResponse({'draw': draw, 'recordsTotal': 0, 'recordsFiltered': 0, 'data': []})

        # Generate the dynamic part of the SQL query based on the columns
        columns_sql = ",\n            ".join([
            f"MAX(CASE WHEN division_name = '{col}' THEN event_titles END) AS \"{col}\"" for col in division_names
        ])

        query = f"""
            SELECT
                generated_date,
                TO_CHAR(generated_date, 'FMMonth DD, YYYY') AS whole_date_start_searchable,
                {columns_sql}
            FROM (
                SELECT
                    generated_date,
                    division_name,
                    STRING_AGG(CONCAT(event_title, '*', id, '*', division_name, '*', unit_name, '*', event_time_start, '*', event_time_end), '|||') AS event_titles
                FROM (
                    SELECT
                        generate_series(whole_date_start::date, whole_date_end::date, '1 day'::interval)::date AS generated_date,
                        division_name,
                        event_title,
                        id,
                        unit_name,
                        event_time_start,
                        event_time_end
                    FROM events_event
                    WHERE office = %s
                      AND display_status = true
                ) AS date_series
                WHERE EXTRACT(YEAR FROM generated_date) = {year}
                GROUP BY generated_date, division_name
            ) AS subquery
            GROUP BY generated_date
        """

        with connection.cursor() as cursor:
            cursor.execute(query, [office])
            result = cursor.fetchall()

        # Convert the result into a list of dictionaries with dynamic keys
        data = []
        for row in result:
            data_row = {
                'whole_date_start': row[0].strftime('%Y-%m-%d'),
                'whole_date_start_searchable': row[1],
                **{division_names[i]: row[i + 2] for i in range(len(division_names))}
            }
            data.append(data_row)

        # Sort the data based on 'whole_date_start'
        order_column_index = int(request.GET.get('order[0][column]', 0))
        order_direction = request.GET.get('order[0][dir]', 'asc')
        data.sort(key=lambda x: x['whole_date_start'], reverse=(order_direction == 'desc'))

        # Apply the search filter to the data
        filtered_data = [
            entry for entry in data if any(
                search_value.lower() in (str(entry[col]).lower() if entry[col] is not None else '') 
                for col in ['whole_date_start_searchable'] + division_names
            )
        ]

        # Generate the columns array with search information
        columns = [
            {'data': 'whole_date_start', 'name': '', 'searchable': True, 'orderable': True, 'search': {} },
            {'data': 'whole_date_start_searchable', 'name': '', 'searchable': True, 'orderable': True, 'search': {'value': search_value} }
        ]
        columns += [{'data': col, 'name': '', 'searchable': True, 'orderable': True, 'search': {} } for col in division_names]

        response_data = {
            'draw': draw,
            'columns': columns,
            'order': [],  # You may need to populate this based on user interaction
            'start': start,
            'length': length,
            'recordsTotal': len(data),
            'recordsFiltered': len(filtered_data),
            'data': filtered_data[start:start + length]
        }

        return JsonResponse(response_data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)



def fetch_events_by_div_ajax_02272024(request):
    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '')

        # Fetch unique division names from the database
        with connection.cursor() as cursor:
            cursor.execute("SELECT DISTINCT division_name FROM divisions_division")
            division_names = [row[0] for row in cursor.fetchall()]

        # Generate the dynamic part of the SQL query based on the columns
        columns_sql = ", ".join([
            f"MAX(CASE WHEN division_name = '{col}' THEN event_titles END) AS {col}" for col in division_names
        ])

        query = f"""
            SELECT
                generated_date,
                TO_CHAR(generated_date, 'FMMonth DD, YYYY') AS whole_date_start_searchable,
                {columns_sql}
            FROM (
                SELECT
                    generated_date,
                    division_name,
                    STRING_AGG(CONCAT(event_title, '*', id, '*', division_name, '*', unit_name, '*', event_time_start, '*', event_time_end), '|||') AS event_titles
                FROM (
                    SELECT
                        generate_series(whole_date_start::date, whole_date_end::date, '1 day'::interval)::date AS generated_date,
                        division_name,
                        event_title,
                        id,
                        unit_name,
                        event_time_start,
                        event_time_end
                    FROM events_event
                    WHERE office = %s
                      AND display_status = true
                ) AS date_series
                GROUP BY generated_date, division_name
            ) AS subquery
            GROUP BY generated_date
            ORDER BY generated_date;
        """

        with connection.cursor() as cursor:
            cursor.execute(query, [request.GET.get('office')])
            result = cursor.fetchall()

        # Sort by 'generated_date' if no specific column is specified
        order_column_index = int(request.GET.get('order[0][column]', 0))
        order_direction = request.GET.get('order[0][dir]', 'asc')

        # Sort by specific column if specified
        if order_direction == 'asc':
            result = sorted(result, key=lambda x: x[order_column_index] if order_column_index < len(result[0]) else x[0])
        else:
            result = sorted(result, key=lambda x: x[order_column_index] if order_column_index < len(result[0]) else x[0], reverse=True)

        # Convert the result into a list of dictionaries with dynamic keys
        data = []
        for row in result:
            data_row = {'whole_date_start': row[0].strftime('%Y-%m-%d')}  # Format as needed

            if isinstance(row[1], str):  # Check if row[1] is a string
                data_row['whole_date_start_searchable'] = row[1]
            else:
                data_row['whole_date_start_searchable'] = row[1].strftime('%B %d, %Y')  # Format whole_date_start_searchable

            data_row.update(zip(division_names, row[2:]))  # Update dynamically using zip
            data.append(data_row)

        # Apply the search filter to the data
        filtered_data = [
            entry for entry in data if any(
                search_value.lower() in (str(entry[col]).lower() if entry[col] is not None else '') 
                for col in ['whole_date_start_searchable'] + division_names
            )
        ]

        # Generate the columns array with search information
        columns = [{'data': 'whole_date_start', 'name': '', 'searchable': True, 'orderable': True, 'search': {} }]
        columns.append({'data': 'whole_date_start_searchable', 'name': '', 'searchable': True, 'orderable': True, 'search': {'value': search_value} })
        columns += [{'data': col, 'name': '', 'searchable': True, 'orderable': True, 'search': {} } for col in division_names]

        response_data = {
            'draw': draw,
            'columns': columns,
            'order': [],  # You may need to populate this based on user interaction
            'start': start,
            'length': length,
            'recordsTotal': len(data),
            'recordsFiltered': len(filtered_data),
            'data': filtered_data[start:start + length]
        }

        return JsonResponse(response_data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def fetch_events_by_div_ajax_ver2(request):
    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '')

        # Fetch unique division names from the database
        with connection.cursor() as cursor:
            cursor.execute("SELECT DISTINCT division_name FROM divisions_division")
            division_names = [row[0] for row in cursor.fetchall()]

        # Generate the dynamic part of the SQL query based on the columns
        columns_sql = ", ".join([
            f"MAX(CASE WHEN division_name = '{col}' THEN event_titles END) AS {col}" for col in division_names
        ])

        query = f"""
            SELECT
                generated_date,
                TO_CHAR(generated_date, 'FMMonth DD, YYYY') AS whole_date_start_searchable,
                {columns_sql}
            FROM (
                SELECT
                    generated_date,
                    division_name,
                    STRING_AGG(CONCAT(event_title, '*', id, '*', division_name, '*', unit_name, '*', event_time_start, '*', event_time_end), '|||') AS event_titles
                FROM (
                    SELECT
                        generate_series(whole_date_start::date, whole_date_end::date, '1 day'::interval)::date AS generated_date,
                        division_name,
                        event_title,
                        id,
                        unit_name,
                        event_time_start,
                        event_time_end
                    FROM events_event
                    WHERE office = %s
                      AND display_status = true
                ) AS date_series
                GROUP BY generated_date, division_name
            ) AS subquery
            GROUP BY generated_date
            ORDER BY generated_date;
        """

        with connection.cursor() as cursor:
            cursor.execute(query, [request.GET.get('office')])
            result = cursor.fetchall()

        # Convert the result into a list of dictionaries with dynamic keys
        data = []
        for row in result:
            data_row = {'whole_date_start': row[0].strftime('%Y-%m-%d')}  # Format as needed

            if isinstance(row[1], str):  # Check if row[1] is a string
                data_row['whole_date_start_searchable'] = row[1]
            else:
                data_row['whole_date_start_searchable'] = row[1].strftime('%B %d, %Y')  # Format whole_date_start_searchable

            data_row.update(zip(division_names, row[2:]))  # Update dynamically using zip
            data.append(data_row)

        # Apply the search filter to the data
        # Apply the search filter to the data
        filtered_data = [
            entry for entry in data if any(
                search_value.lower() in (str(entry[col]).lower() if entry[col] is not None else '') 
                for col in ['whole_date_start_searchable'] + division_names
            )
        ]


        # Generate the columns array with search information
        # Generate the columns array with search information
        columns = [{'data': 'whole_date_start', 'name': '', 'searchable': True, 'orderable': True, 'search': {} }]
        columns.append({'data': 'whole_date_start_searchable', 'name': '', 'searchable': True, 'orderable': True, 'search': {'value': search_value} })
        columns += [{'data': col, 'name': '', 'searchable': True, 'orderable': True, 'search': {} } for col in division_names]

        response_data = {
            'draw': draw,
            'columns': columns,
            'order': [],  # You may need to populate this based on user interaction
            'start': start,
            'length': length,
            'recordsTotal': len(data),
            'recordsFiltered': len(filtered_data),
            'data': filtered_data[start:start + length]
        }

        return JsonResponse(response_data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

    # get events by unit
def fetch_events_by_unit_ajax(request):
    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '')
        #office_txt = request.GET.get('office', '')

        # Define the columns you want to search on
        # columns = ['whole_date_start_searchable']
        columns = ['whole_date_start_searchable', 'NC', 'EDU', 'TPU', 'CARP', 'GAD', 'PLANNING']

        query = """
        SELECT whole_date_start,
            whole_date_start_searchable, 
            MAX(CASE WHEN unit = 'NC' THEN event_titles END) AS NC,
            MAX(CASE WHEN unit = 'EDU' THEN event_titles END) AS EDU,
            MAX(CASE WHEN unit = 'TPU' THEN event_titles END) AS TPU,
            MAX(CASE WHEN unit = 'CARP' THEN event_titles END) AS CARP,
            MAX(CASE WHEN unit = 'GAD' THEN event_titles END) AS GAD,
            MAX(CASE WHEN unit = 'PLANNING' THEN event_titles END) AS PLANNING
        FROM (
            SELECT whole_date_start,
                whole_date_start_searchable,
                unit,
                STRING_AGG(event_title, ', ') AS event_titles, division_name, office
            FROM events_event
            WHERE division_name = %s AND office = %s
            GROUP BY whole_date_start, whole_date_start_searchable, unit, division_name, office
        ) AS subquery
        GROUP BY whole_date_start, whole_date_start_searchable
        ORDER BY 1;
        """

        with connection.cursor() as cursor:
            cursor.execute(query, [request.GET.get('division'), request.GET.get('office')])
            result = cursor.fetchall()

        # Convert the result into a list of dictionaries
        data = []
        for row in result:
            data.append({
                'whole_date_start': row[0].strftime('%Y-%m-%d'),  # Format as needed
                'whole_date_start_searchable': row[1],
                'NC': row[2],
                'EDU': row[3],
                'TPU': row[4],
                'CARP': row[5],
                'GAD': row[6],
                'PLANNING': row[7]
                # Add more fields as needed
            })
            
        # Apply the search filter to the data
        # filtered_data = [entry for entry in data if any(search_value.lower() in entry[col].lower() for col in columns)]
        filtered_data = [
            entry for entry in data if any(
                search_value.lower() in (str(entry[col]).lower() if entry[col] is not None else '') for col in columns
            )
        ]

        response_data = {
            'draw': draw,
            'recordsTotal': len(data),
            'recordsFiltered': len(filtered_data),
            'data': filtered_data[start:start + length]
        }

        return JsonResponse(response_data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
    # method to download files from the media folder
def download_file(request, id):
    event = get_object_or_404(Event, pk=id)
    file_path = event.file_attachment.path

    # Check if the file_path exists before attempting to download
    if os.path.exists(file_path):
        try:
            return FileResponse(open(file_path, 'rb'), content_type='application/octet-stream')
        except FileNotFoundError:
            return HttpResponse('File not found', status=404)
    else:
        return HttpResponse('No file associated', status=404)
    

# method to display tooltips
def tooltips(request):
    return render(request, 'events/tooltip.html')

@csrf_exempt
def mark_event_false(request):
    
    if request.method == 'POST':

        eventPID = request.POST['event-primaryID']
        existing_event = Event.objects.filter(id=eventPID).first()
        existing_event.display_status = False
       # Save the update display status
        existing_event.save()

        return JsonResponse({'message': 'True'})
    else:
        return JsonResponse({'message': 'False'})

        # # Debugging: Print the event ID to verify
        # print("Event ID:", eventPID)
        # return JsonResponse({'message': eventPID})


# ---------------------------------------------------------------------------
# Server-side Excel / PDF export helpers
# ---------------------------------------------------------------------------

def _parse_event_cell(cell_data):
    """Parse '|||' separated events from a cell into list of dicts."""
    if not cell_data:
        return []
    events = []
    for part in cell_data.split('|||'):
        part = part.strip()
        if not part:
            continue
        segs = part.split('*')
        events.append({
            'title': segs[0] if len(segs) > 0 else '',
            'division': segs[2] if len(segs) > 2 else '',
            'unit': segs[3] if len(segs) > 3 else '',
            'time_start': segs[4] if len(segs) > 4 else '',
            'time_end': segs[5] if len(segs) > 5 else '',
        })
    return events


def _format_cell_text(cell_data):
    """Convert raw cell data to clean plain text for export."""
    events = _parse_event_cell(cell_data)
    if not events:
        return ''
    lines = []
    for ev in events:
        line = f"\u2022 {ev['title']}"
        meta = []
        if ev['division']:
            meta.append(ev['division'])
        if ev['unit']:
            meta.append(ev['unit'])
        if ev['time_start']:
            t = ev['time_start']
            if ev['time_end']:
                t += f" \u2013 {ev['time_end']}"
            meta.append(t)
        if meta:
            line += f"\n  ({' | '.join(meta)})"
        lines.append(line)
    return '\n'.join(lines)


def _get_office_export_data(year):
    """Return (columns, data) for office-level export."""
    offices = list(Office.objects.order_by('id').values_list('office_initials', flat=True))
    case_statements = ',\n            '.join([
        f"MAX(CASE WHEN office = '{o}' THEN event_titles END) AS \"{o}\""
        for o in offices
    ])
    query = f"""
        SELECT
            DISTINCT generated_date,
            TO_CHAR(generated_date, 'FMMonth DD, YYYY') AS whole_date_start_searchable,
            {case_statements}
        FROM (
            SELECT
                generated_date,
                office,
                STRING_AGG(CONCAT(event_title, '*', id, '*', division_name, '*', unit_name, '*', event_time_start, '*', event_time_end), '|||') AS event_titles
            FROM (
                SELECT
                    generate_series(whole_date_start::date, whole_date_end::date, '1 day'::interval)::date AS generated_date,
                    office,
                    event_title,
                    id,
                    division_name,
                    unit_name,
                    event_time_start,
                    event_time_end
                FROM events_event
                WHERE display_status = true
            ) AS date_series
            WHERE EXTRACT(YEAR FROM generated_date) = {year}
            GROUP BY generated_date, office
        ) AS subquery
        GROUP BY generated_date
        ORDER BY generated_date ASC;
    """
    with connection.cursor() as cursor:
        cursor.execute(query)
        rows = cursor.fetchall()
    data = []
    for row in rows:
        entry = {
            'whole_date_start_searchable': row[1],
        }
        for i, office in enumerate(offices):
            entry[office] = row[2 + i]
        data.append(entry)
    return offices, data


def _get_div_export_data(year, office):
    """Return (columns, data) for division-level export."""
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT dd.division_name
            FROM divisions_division dd
            INNER JOIN tbl_offices o ON dd.fk_office_id = o.id
            WHERE o.office_initials = %s
            ORDER BY dd.id
        """, [office])
        division_names = [row[0] for row in cursor.fetchall()]

    if not division_names:
        return [], []

    columns_sql = ',\n            '.join([
        f"MAX(CASE WHEN division_name = '{col}' THEN event_titles END) AS \"{col}\""
        for col in division_names
    ])
    query = f"""
        SELECT
            generated_date,
            TO_CHAR(generated_date, 'FMMonth DD, YYYY') AS whole_date_start_searchable,
            {columns_sql}
        FROM (
            SELECT
                generated_date,
                division_name,
                STRING_AGG(CONCAT(event_title, '*', id, '*', division_name, '*', unit_name, '*', event_time_start, '*', event_time_end), '|||') AS event_titles
            FROM (
                SELECT
                    generate_series(whole_date_start::date, whole_date_end::date, '1 day'::interval)::date AS generated_date,
                    division_name,
                    event_title,
                    id,
                    unit_name,
                    event_time_start,
                    event_time_end
                FROM events_event
                WHERE office = %s AND display_status = true
            ) AS date_series
            WHERE EXTRACT(YEAR FROM generated_date) = {year}
            GROUP BY generated_date, division_name
        ) AS subquery
        GROUP BY generated_date
        ORDER BY generated_date ASC;
    """
    with connection.cursor() as cursor:
        cursor.execute(query, [office])
        rows = cursor.fetchall()
    data = []
    for row in rows:
        entry = {'whole_date_start_searchable': row[1]}
        for i, div in enumerate(division_names):
            entry[div] = row[2 + i]
        data.append(entry)
    return division_names, data


# ---------------------------------------------------------------------------
# Excel export view
# ---------------------------------------------------------------------------

@login_required
def export_events_excel(request):
    year = int(request.GET.get('year', timezone.now().year))
    office = request.GET.get('office', '').strip()

    if office:
        columns, data = _get_div_export_data(year, office)
        title = f"DTI CARAGA CALENDAR OF EVENTS BY DIVISION \u2013 {office.upper()} \u2013 {year}"
        filename = f"dti_events_{office}_{year}.xlsx"
    else:
        columns, data = _get_office_export_data(year)
        title = f"DTI CARAGA CALENDAR OF EVENTS \u2013 {year}"
        filename = f"dti_events_{year}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = f"Events {year}"

    # Styles
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    title_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    header_font = Font(name='Calibri', color="F1F5F9", bold=True, size=10)
    title_font = Font(name='Calibri', bold=True, size=13, color="1E293B")
    date_font = Font(name='Calibri', bold=True, size=9, color="1E293B")
    cell_font = Font(name='Calibri', size=9, color="334155")
    thin = Side(style='thin', color='94A3B8')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    total_cols = len(columns) + 1  # +1 for date column

    # Row 1 — Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.border = thin_border
    ws.row_dimensions[1].height = 28

    # Row 2 — Header
    ws.cell(row=2, column=1, value='DATE').fill = header_fill
    ws.cell(row=2, column=1).font = header_font
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=2, column=1).border = thin_border
    ws.row_dimensions[2].height = 22

    for col_idx, col_name in enumerate(columns, start=2):
        cell = ws.cell(row=2, column=col_idx, value=col_name.upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Data rows
    for row_idx, entry in enumerate(data, start=3):
        fill = alt_fill if row_idx % 2 == 0 else None

        date_cell = ws.cell(row=row_idx, column=1, value=entry['whole_date_start_searchable'])
        date_cell.font = date_font
        date_cell.alignment = Alignment(horizontal='center', vertical='top')
        date_cell.border = thin_border
        if fill:
            date_cell.fill = fill

        for col_idx, col_name in enumerate(columns, start=2):
            text = _format_cell_text(entry.get(col_name))
            cell = ws.cell(row=row_idx, column=col_idx, value=text if text else None)
            cell.font = cell_font
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border
            if fill:
                cell.fill = fill

    # Column widths
    ws.column_dimensions['A'].width = 22
    for col_idx in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 32

    # Freeze panes — freeze header rows
    ws.freeze_panes = 'A3'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ---------------------------------------------------------------------------
# PDF export view
# ---------------------------------------------------------------------------

@login_required
def export_events_pdf(request):
    year = int(request.GET.get('year', timezone.now().year))
    office = request.GET.get('office', '').strip()

    if office:
        columns, data = _get_div_export_data(year, office)
        title = f"DTI CARAGA CALENDAR OF EVENTS BY DIVISION \u2013 {office.upper()} \u2013 {year}"
        filename = f"dti_events_{office}_{year}.pdf"
    else:
        columns, data = _get_office_export_data(year)
        title = f"DTI CARAGA CALENDAR OF EVENTS \u2013 {year}"
        filename = f"dti_events_{year}.pdf"

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=1.2 * cm, rightMargin=1.2 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'ExportTitle', parent=styles['Normal'],
        fontSize=13, fontName='Helvetica-Bold',
        alignment=TA_CENTER, spaceAfter=10,
        textColor=colors.HexColor('#1E293B'),
    )
    header_style = ParagraphStyle(
        'ColHeader', parent=styles['Normal'],
        fontSize=8, fontName='Helvetica-Bold',
        textColor=colors.white, alignment=TA_CENTER,
    )
    date_style = ParagraphStyle(
        'DateCell', parent=styles['Normal'],
        fontSize=8, fontName='Helvetica-Bold',
        textColor=colors.HexColor('#1E293B'), alignment=TA_CENTER,
    )
    cell_style = ParagraphStyle(
        'EventCell', parent=styles['Normal'],
        fontSize=7, fontName='Helvetica',
        textColor=colors.HexColor('#334155'), leading=10,
    )

    header_color = colors.HexColor('#1E293B')
    alt_color = colors.HexColor('#F8FAFC')
    border_color = colors.HexColor('#94A3B8')

    # Build table data
    table_data = [
        [Paragraph('DATE', header_style)] +
        [Paragraph(col.upper(), header_style) for col in columns]
    ]
    for entry in data:
        row = [Paragraph(entry['whole_date_start_searchable'], date_style)]
        for col_name in columns:
            text = _format_cell_text(entry.get(col_name))
            safe_text = text.replace('\n', '<br/>').replace('\u2022', '&#8226;') if text else ''
            row.append(Paragraph(safe_text, cell_style))
        table_data.append(row)

    # Column widths — date col fixed, rest split evenly
    page_width = landscape(A4)[0] - 2.4 * cm
    date_col_w = 3.8 * cm
    other_col_w = (page_width - date_col_w) / len(columns) if columns else page_width
    col_widths = [date_col_w] + [other_col_w] * len(columns)

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), header_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, border_color),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, alt_color]),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))

    def add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(colors.HexColor('#94A3B8'))
        canvas.drawRightString(
            landscape(A4)[0] - 1.2 * cm,
            0.8 * cm,
            f"Page {doc.page}"
        )
        canvas.restoreState()

    doc.build(
        [Paragraph(title, title_style), Spacer(1, 0.4 * cm), tbl],
        onFirstPage=add_page_number,
        onLaterPages=add_page_number,
    )
    return response
    


    


        






    

